from flask import Blueprint, request, jsonify, send_file, current_app, session, redirect, url_for, flash, render_template
from models import db, Employee, Entry
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from utils.data_processing import get_weekly_evolution_data
from utils.email_utils import send_confirmation_email
import tempfile
from utils.calculations import calculate_weekly_progress, get_current_week, get_week_from_date
import io
import zipfile
from calendar import monthrange
import time

api_bp = Blueprint('api', __name__)

# Cache simples em memória - OTIMIZADO
_cache = {}
_cache_timeout = 600  # 10 minutos (aumentado)
_cache_stats = {'hits': 0, 'misses': 0}

# Pré-carregamento de dados críticos
_critical_data = {}
_last_preload = 0
_preload_interval = 300  # 5 minutos

def preload_critical_data():
    """Pré-carrega dados críticos para eliminar lag"""
    global _critical_data, _last_preload
    
    current_time = time.time()
    if current_time - _last_preload < _preload_interval:
        return _critical_data
    
    try:
        current_app.logger.info("🔄 Pré-carregando dados críticos...")
        
        # Buscar todos os dados de uma vez
        employees = Employee.query.all()
        all_entries = Entry.query.all()
        
        # Processar dados críticos
        employee_data = {}
        weekly_data = {}
        monthly_data = {}
        
        for employee in employees:
            employee_data[employee.id] = {
                'id': employee.id,
                'name': employee.real_name,
                'username': employee.username,
                'role': employee.display_role,
                'weekly_goal': employee.weekly_goal
            }
        
        # Calcular dados semanais
        from utils.calculations import get_week_dates
        week_dates = []
        for week in range(1, 6):
            start_date, end_date = get_week_dates(str(week))
            week_dates.append((start_date, end_date))
        
        # Processar todos os registros
        for entry in all_entries:
            entry_date = entry.date
            if hasattr(entry_date, 'strftime'):
                entry_date_str = entry_date.strftime('%Y-%m-%d')
            else:
                entry_date_str = str(entry_date)[:10]
            
            # Semanal
            for week_idx, (start_date, end_date) in enumerate(week_dates):
                if start_date <= entry_date_str <= end_date:
                    if entry.employee_id not in weekly_data:
                        weekly_data[entry.employee_id] = [0] * 5
                    weekly_data[entry.employee_id][week_idx] += entry.points
                    break
        
        # Calcular progresso para cada funcionário
        for employee in employees:
            if employee.id in weekly_data:
                weekly_points = sum(weekly_data[employee.id])
                weekly_goal = employee.weekly_goal
                progress_percentage = (weekly_points / weekly_goal * 100) if weekly_goal > 0 else 0
                remaining_points = max(0, weekly_goal - weekly_points)
                
                employee_data[employee.id].update({
                    'weekly_points': weekly_points,
                    'progress_percentage': progress_percentage,
                    'remaining_points': remaining_points
                })
        
        _critical_data = {
            'employees': employee_data,
            'weekly_data': weekly_data,
            'last_update': current_time
        }
        
        _last_preload = current_time
        current_app.logger.info("✅ Dados críticos pré-carregados com sucesso")
        
        return _critical_data
        
    except Exception as e:
        current_app.logger.error(f"Erro ao pré-carregar dados: {str(e)}")
        return _critical_data

def get_cached_data(key):
    """Obtém dados do cache se ainda válidos"""
    if key in _cache:
        data, timestamp = _cache[key]
        if time.time() - timestamp < _cache_timeout:
            _cache_stats['hits'] += 1
            return data
    _cache_stats['misses'] += 1
    return None

def set_cached_data(key, data):
    """Armazena dados no cache"""
    _cache[key] = (data, time.time())

def clear_cache():
    """Limpa todo o cache"""
    _cache.clear()
    _cache_stats['hits'] = 0
    _cache_stats['misses'] = 0

def get_cache_stats():
    """Retorna estatísticas do cache"""
    total = _cache_stats['hits'] + _cache_stats['misses']
    hit_rate = (_cache_stats['hits'] / total * 100) if total > 0 else 0
    return {
        'hit_rate': hit_rate,
        'hits': _cache_stats['hits'],
        'misses': _cache_stats['misses'],
        'total': total
    }

@api_bp.route('/register_points', methods=['POST'])
def register_points():
    if 'role' not in session or session['role'] != 'employee':
        return redirect(url_for('auth.index'))
    
    try:
        employee_id = session['user_id']
        date = request.form['date']
        refinery = request.form['refinery']
        points = int(request.form['points'])
        observations = request.form.get('observations', '')
        
        # Adicionar horário atual à data (igual ao app_antigo)
        from datetime import datetime
        from utils.helpers import timezone
        current_time = datetime.now(timezone).strftime('%H:%M:%S')
        date_with_time = f"{date} {current_time}"
        
        # Permitir múltiplos registros na mesma data (não checamos duplicidade)
        new_entry = Entry(
            employee_id=employee_id,
            date=date_with_time,
            refinery=refinery,
            points=points,
            observations=observations
        )

        db.session.add(new_entry)
        db.session.commit()
        
        # Invalidar cache relacionado a este funcionário
        # cache.delete_memoized(get_employees_summary_cached)
        # cache.delete_memoized(get_weekly_data_cached)
        # cache.delete_memoized(get_monthly_data_cached)
        # cache.delete_memoized(get_daily_data_cached)
        
        # Buscar informações do funcionário para o email
        employee = Employee.query.get(employee_id)
        if employee:
            # Enviar email de confirmação
            send_confirmation_email(
                employee.real_name,
                date,
                points,
                refinery,
                observations
            )
        
        flash('Ponto registrado com sucesso! Até amanhã!', 'success')
        
        return redirect(url_for('dashboard.employee_dashboard_enhanced'))
    except Exception as e:
        current_app.logger.error(f"Erro ao registrar pontos: {str(e)}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500

@api_bp.route('/edit_entry/<int:entry_id>', methods=['GET', 'POST'])
def edit_entry(entry_id):
    if 'role' not in session:
        return redirect(url_for('auth.index'))
    
    entry = Entry.query.get_or_404(entry_id)
    
    # Verificar permissões
    if session['role'] == 'employee' and entry.employee_id != session['user_id']:
        flash('Você não tem permissão para editar este registro!', 'error')
        return redirect(url_for('dashboard.employee_dashboard_enhanced'))
    
    if request.method == 'POST':
        try:
            # Combina data e hora (se fornecido)
            date_part = request.form['date']
            time_part = request.form.get('time', '').strip()
            if time_part and len(time_part)==5:  # HH:MM
                time_part += ':00'
            entry.date = f"{date_part} {time_part}" if time_part else date_part
            entry.refinery = request.form['refinery']
            entry.points = int(request.form['points'])
            entry.observations = request.form.get('observations', '')
            
            db.session.commit()
            
            # Invalidar cache relacionado
            # cache.delete_memoized(get_employees_summary_cached)
            # cache.delete_memoized(get_weekly_data_cached)
            # cache.delete_memoized(get_monthly_data_cached)
            # cache.delete_memoized(get_daily_data_cached)
            
            flash('Registro atualizado com sucesso!', 'success')
            
            # Redirecionar baseado no papel do usuário
            if session['role'] == 'ceo':
                week = request.form.get('week')
                employee_id = request.form.get('employee_id')
                return redirect(url_for('dashboard.ceo_dashboard_enhanced', week=week, employee_id=employee_id))
            else:
                return redirect(url_for('dashboard.employee_dashboard_enhanced'))
        except Exception as e:
            current_app.logger.error(f"Erro ao editar entrada: {str(e)}")
            return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500
    else:
        # Método GET: renderizar formulário de edição
        # Extrair partes de data e hora para pré-preencher inputs
        if hasattr(entry.date, 'strftime'):
            iso_date = entry.date.strftime('%Y-%m-%d')
            iso_time = entry.date.strftime('%H:%M')
        else:
            # Se entry.date é string
            try:
                date_obj = datetime.strptime(entry.date, '%Y-%m-%d %H:%M:%S')
                iso_date = date_obj.strftime('%Y-%m-%d')
                iso_time = date_obj.strftime('%H:%M')
            except:
                iso_date = entry.date
                iso_time = ''

        return render_template('edit_entry.html', entry=entry, iso_date=iso_date, iso_time=iso_time)

@api_bp.route('/api/delete_entry/<int:entry_id>', methods=['POST'])
@api_bp.route('/delete_entry/<int:entry_id>', methods=['POST'])
def delete_entry(entry_id):
    if 'role' not in session:
        return redirect(url_for('auth.index'))
    
    entry = Entry.query.get_or_404(entry_id)
    
    # Verificar permissões
    if session['role'] == 'employee' and entry.employee_id != session['user_id']:
        flash('Você não tem permissão para deletar este registro!', 'error')
        return redirect(url_for('dashboard.employee_dashboard_enhanced'))
    
    try:
        db.session.delete(entry)
        db.session.commit()
        
        # Invalidar cache relacionado
        # cache.delete_memoized(get_employees_summary_cached)
        # cache.delete_memoized(get_weekly_data_cached)
        # cache.delete_memoized(get_monthly_data_cached)
        # cache.delete_memoized(get_daily_data_cached)
        
        flash('Registro deletado com sucesso!', 'success')
        return redirect(url_for('dashboard.employee_dashboard_enhanced'))
    except Exception as e:
        current_app.logger.error(f"Erro ao deletar entrada: {str(e)}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500

@api_bp.route('/api/delete_all_entries', methods=['POST'])
@api_bp.route('/delete_all_entries', methods=['POST'])
def delete_all_entries():
    if 'role' not in session or session['role'] != 'ceo':
        flash('Acesso negado!', 'error')
        return redirect(url_for('auth.index'))
    
    try:
        Entry.query.delete()
        db.session.commit()
        
        # Invalidar todo o cache
        # cache.clear()
        
        flash('Todos os registros foram deletados!', 'success')
        return redirect(url_for('dashboard.ceo_dashboard_enhanced'))
    except Exception as e:
        current_app.logger.error(f"Erro ao deletar todas as entradas: {str(e)}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500

@api_bp.route('/api/export')
@api_bp.route('/export')
def export():
    if 'role' not in session:
        return redirect(url_for('auth.index'))
    
    try:
        # Buscar todos os funcionários
        employees = Employee.query.all()

        if not employees:
            flash('Nenhum funcionário encontrado para exportar!', 'warning')
            return redirect(url_for('dashboard.ceo_dashboard_enhanced'))
        
        # Criar ZIP com arquivos separados por funcionário
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for employee in employees:
                # Buscar registros do funcionário
                entries = Entry.query.filter_by(employee_id=employee.id).order_by(Entry.date.desc()).all()
                
                if entries:
                    # Criar workbook para o funcionário
                    wb = Workbook()
                    ws = wb.active
                    ws.title = employee.real_name

                    # Estilos
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    # Cabeçalhos
                    headers = ['Data', 'Refinaria', 'Pontos', 'Observações']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")

                    # Dados
                    total_points = 0
                    for row, entry in enumerate(entries, 2):
                        ws.cell(row=row, column=1, value=entry.date)
                        ws.cell(row=row, column=2, value=entry.refinery)
                        ws.cell(row=row, column=3, value=entry.points)
                        ws.cell(row=row, column=4, value=entry.observations or '')
                        total_points += entry.points
                    
                    # Linha de total
                    remaining_monthly = max(0, employee.monthly_goal - total_points)
                    
                    total_row = len(entries) + 2
                    ws.cell(row=total_row, column=1, value='Total')
                    ws.cell(row=total_row, column=2, value='')
                    ws.cell(row=total_row, column=3, value=total_points)
                    ws.cell(row=total_row, column=4, value=f'Restante mensal: {remaining_monthly}')

                    # Ajustar largura das colunas
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Salvar arquivo do funcionário no ZIP
                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    zip_file.writestr(f"{employee.real_name}.xlsx", excel_buffer.read())
        
        zip_buffer.seek(0)
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name='relatorios_funcionarios.zip',
            mimetype='application/zip'
        )
        
    except Exception as e:
        current_app.logger.error(f"Erro ao exportar dados: {str(e)}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500

@api_bp.route('/employees')
@api_bp.route('/api/employees')
def get_employees_summary():
    """Endpoint para obter resumo dos funcionários - ULTRA OTIMIZADO"""
    try:
        # Usar dados pré-carregados
        critical_data = preload_critical_data()
        
        employees_data = []
        for employee_data in critical_data['employees'].values():
            employees_data.append({
                'id': employee_data['id'],
                'name': employee_data['name'],
                'username': employee_data['username'],
                'role': employee_data['role'],
                'weekly_points': employee_data.get('weekly_points', 0),
                'weekly_goal': employee_data['weekly_goal'],
                'progress_percentage': employee_data.get('progress_percentage', 0),
                'remaining_points': employee_data.get('remaining_points', 0)
            })
        
        return jsonify({'employees': employees_data})
        
    except Exception as e:
        current_app.logger.error(f"Erro ao obter resumo dos funcionários: {str(e)}")
        return jsonify({'employees': []})

@api_bp.route('/team-progress')
def get_team_progress():
    """Endpoint para obter progresso da equipe - ULTRA OTIMIZADO"""
    try:
        # Usar dados pré-carregados
        critical_data = preload_critical_data()
        
        team_data = []
        total_points = 0
        total_goal = 0
        
        for employee_data in critical_data['employees'].values():
            weekly_points = employee_data.get('weekly_points', 0)
            weekly_goal = employee_data['weekly_goal']
            progress_percentage = employee_data.get('progress_percentage', 0)
            remaining_points = employee_data.get('remaining_points', 0)
            
            employee_info = {
                'id': employee_data['id'],
                'name': employee_data['name'],
                'weekly_points': weekly_points,
                'weekly_goal': weekly_goal,
                'progress_percentage': progress_percentage,
                'remaining_points': remaining_points
            }
            
            team_data.append(employee_info)
            total_points += weekly_points
            total_goal += weekly_goal
    
        return jsonify({
            'employees': team_data,
            'total_points': total_points,
            'total_goal': total_goal,
            'team_progress_percentage': (total_points / total_goal * 100) if total_goal > 0 else 0
        })
        
    except Exception as e:
        current_app.logger.error(f"Erro ao calcular progresso da equipe: {str(e)}")
        return jsonify({'employees': [], 'total_points': 0, 'total_goal': 0, 'team_progress_percentage': 0})

@api_bp.route('/api/weekly_data')
def api_weekly_data():
    """Endpoint para dados semanais - COM CACHE"""
    try:
        # Verificar cache primeiro
        cache_key = 'weekly_data'
        cached_data = get_cached_data(cache_key)
        if cached_data:
            return jsonify(cached_data)
        
        # Se não estiver em cache, buscar dados
        from utils.data_processing import get_weekly_progress_data
        weekly_data = get_weekly_progress_data()
        
        # Armazenar no cache
        set_cached_data(cache_key, weekly_data)
        
        return jsonify(weekly_data)
    except Exception as e:
        current_app.logger.error(f"Erro ao obter dados semanais: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@api_bp.route('/api/monthly_data')
def api_monthly_data():
    """Endpoint para dados mensais - COM CACHE"""
    try:
        # Verificar cache primeiro
        cache_key = 'monthly_data'
        cached_data = get_cached_data(cache_key)
        if cached_data:
            return jsonify(cached_data)
        
        # Se não estiver em cache, buscar dados
        from utils.data_processing import get_monthly_evolution_data
        monthly_data = get_monthly_evolution_data()
        
        # Armazenar no cache
        set_cached_data(cache_key, monthly_data)
        
        return jsonify(monthly_data)
    except Exception as e:
        current_app.logger.error(f"Erro ao obter dados mensais: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@api_bp.route('/api/daily_data')
def api_daily_data():
    """Endpoint para dados diários - COM CACHE"""
    try:
        # Verificar cache primeiro
        cache_key = 'daily_data'
        cached_data = get_cached_data(cache_key)
        if cached_data:
            return jsonify(cached_data)
        
        # Se não estiver em cache, buscar dados
        from utils.data_processing import get_daily_data
        daily_data = get_daily_data()
        
        # Armazenar no cache
        set_cached_data(cache_key, daily_data)
        
        return jsonify(daily_data)
    except Exception as e:
        current_app.logger.error(f"Erro ao obter dados diários: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@api_bp.route('/api/entries')
def api_entries():
    """Endpoint para obter registros"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        employee_id = request.args.get('employee_id', type=int)
        week = request.args.get('week', type=str)

        current_app.logger.info(f"🔍 DEBUG MASSIVO: Iniciando api_entries")
        current_app.logger.info(f"🔍 DEBUG MASSIVO: page={page}, per_page={per_page}, employee_id={employee_id}, week={week}")

        query = Entry.query.join(Employee)
        current_app.logger.info(f"🔍 DEBUG MASSIVO: Query criada com JOIN Employee")
        
        if employee_id:
            query = query.filter(Entry.employee_id == employee_id)
            current_app.logger.info(f"🔍 DEBUG MASSIVO: Filtro por employee_id={employee_id} aplicado")

        # ✅ ADICIONAR FILTRO DE SEMANA
        if week and week != '':
            current_app.logger.info(f"🔍 DEBUG MASSIVO: Aplicando filtro de semana: {week}")
            from utils.calculations import get_week_dates
            start_date, end_date = get_week_dates(week)
            current_app.logger.info(f"🔍 DEBUG MASSIVO: Semana {week} - De {start_date} até {end_date}")
            
            # Filtrar por data da semana
            query = query.filter(Entry.date >= start_date, Entry.date <= end_date)
            current_app.logger.info(f"🔍 DEBUG MASSIVO: Filtro de data aplicado: {start_date} <= date <= {end_date}")

        # Paginação
        pagination = query.order_by(Entry.date.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        current_app.logger.info(f"🔍 DEBUG MASSIVO: Paginação criada - total={pagination.total}, pages={pagination.pages}")
        
        entries_data = []
        current_app.logger.info(f"🔍 DEBUG MASSIVO: Processando {len(pagination.items)} registros")
        
        for i, entry in enumerate(pagination.items):
            current_app.logger.info(f"🔍 DEBUG MASSIVO: Processando registro {i+1}/{len(pagination.items)}")
            current_app.logger.info(f"🔍 DEBUG MASSIVO: Entry ID: {entry.id}")
            current_app.logger.info(f"🔍 DEBUG MASSIVO: Entry employee_id: {entry.employee_id}")
            
            # DEBUG MASSIVO: Verificar se o employee existe
            current_app.logger.info(f"🔍 DEBUG MASSIVO: hasattr(entry, 'employee'): {hasattr(entry, 'employee')}")
            current_app.logger.info(f"🔍 DEBUG MASSIVO: entry.employee: {entry.employee}")
            
            if hasattr(entry, 'employee') and entry.employee:
                current_app.logger.info(f"🔍 DEBUG MASSIVO: Employee encontrado!")
                current_app.logger.info(f"🔍 DEBUG MASSIVO: Employee ID: {entry.employee.id}")
                current_app.logger.info(f"🔍 DEBUG MASSIVO: Employee real_name: {entry.employee.real_name}")
                current_app.logger.info(f"🔍 DEBUG MASSIVO: Employee username: {entry.employee.username}")
                
                employee_name = entry.employee.real_name
                current_app.logger.info(f"🔍 DEBUG MASSIVO: Employee name definido como: '{employee_name}'")
            else:
                current_app.logger.warning(f"🔍 DEBUG MASSIVO: Employee NÃO encontrado para entry {entry.id}")
                current_app.logger.warning(f"🔍 DEBUG MASSIVO: Tentando buscar employee manualmente...")
                
                # Tentar buscar employee manualmente
                try:
                    employee = Employee.query.get(entry.employee_id)
                    if employee:
                        current_app.logger.info(f"🔍 DEBUG MASSIVO: Employee encontrado manualmente: {employee.real_name}")
                        employee_name = employee.real_name
                    else:
                        current_app.logger.error(f"🔍 DEBUG MASSIVO: Employee com ID {entry.employee_id} NÃO existe no banco!")
                        employee_name = "Funcionário Desconhecido"
                except Exception as e:
                    current_app.logger.error(f"🔍 DEBUG MASSIVO: Erro ao buscar employee: {str(e)}")
                    employee_name = "Erro ao buscar funcionário"
            
            current_app.logger.info(f"🔍 DEBUG MASSIVO: Employee name final: '{employee_name}'")
            
            entry_data = {
                'id': entry.id,
                'employee_name': employee_name,
                'date': entry.date,
                'refinery': entry.refinery,
                'points': entry.points,
                'observations': entry.observations
            }
            
            current_app.logger.info(f"🔍 DEBUG MASSIVO: Entry data criado: {entry_data}")
            entries_data.append(entry_data)
        
        current_app.logger.info(f"🔍 DEBUG MASSIVO: Total de entries_data: {len(entries_data)}")
        
        response_data = {
            'entries': entries_data,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            }
        }
        
        current_app.logger.info(f"🔍 DEBUG MASSIVO: Response data criado com {len(entries_data)} entries")
        
        return jsonify(response_data)
        
    except Exception as e:
        current_app.logger.error(f"🔍 DEBUG MASSIVO: Erro ao obter registros: {str(e)}")
        current_app.logger.error(f"🔍 DEBUG MASSIVO: Stack trace: {e.__traceback__}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@api_bp.route('/api/cache/stats')
def cache_stats():
    """Endpoint para ver estatísticas do cache"""
    try:
        stats = get_cache_stats()
        return jsonify({
            'cache_stats': stats,
            'cache_size': len(_cache),
            'cache_timeout': _cache_timeout
        })
    except Exception as e:
        current_app.logger.error(f"Erro ao obter estatísticas do cache: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@api_bp.route('/api/cache/clear', methods=['POST'])
def clear_cache_endpoint():
    """Endpoint para limpar o cache"""
    try:
        clear_cache()
        return jsonify({'success': True, 'message': 'Cache limpo com sucesso'})
    except Exception as e:
        current_app.logger.error(f"Erro ao limpar cache: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500